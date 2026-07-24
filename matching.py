"""
模板匹配回填引擎 (Template Matching Engine)。

业务场景：用户上传 xlsx 模板（含金额列和发票号回填列），系统从库内发票按金额匹配回填发票号。
  - 第一轮：一对一金额匹配（精确到分），相同金额时用发票备注做模糊校验打分
  - 第二轮：多对一凑票（模板多行金额相加 = 一张发票金额）

依赖 db.get_invoices_for_matching(buyer, date_from, date_to) 取候选发票。
"""
import io
import time
from datetime import datetime, date, timedelta
from itertools import combinations

from openpyxl import load_workbook

import db


# ----------------------------------------------------------- 配置对象
class MatchConfig:
    """匹配配置。支持属性访问；也可直接传 dict 给各匹配函数（用 _cfg_get 兼容）。"""

    def __init__(self, date_range_days=30, overwrite=False,
                 max_combo_size=5, group_time_budget=2.0):
        self.date_range_days = date_range_days
        self.overwrite = overwrite
        self.max_combo_size = max_combo_size
        self.group_time_budget = group_time_budget


def _cfg_get(config, name, default):
    """从 config（对象或 dict 或 None）读取字段，缺省返回 default。"""
    if config is None:
        return default
    if isinstance(config, dict):
        return config.get(name, default)
    return getattr(config, name, default)


# ----------------------------------------------------------- 列识别关键词
_KW_AMOUNT = ("金额", "支出")
_KW_INVOICE_NO = ("发票号", "开票号")
_KW_DATE = ("日期",)
_KW_BUYER = ("收款", "买方", "单位")
_KW_MERCHANT = ("商社", "销售方")


def _match_keywords(header_text, keywords):
    """表头单元格文本是否命中任一关键词。"""
    if not header_text:
        return False
    s = str(header_text)
    return any(kw in s for kw in keywords)


def _find_columns(headers):
    """扫描表头，识别 5 类列。返回 (columns_dict, ambiguous)。
    列索引为 0-based。某类列出现多个时取首个并标记 ambiguous。"""
    amount_cols, invoice_no_cols, date_cols, buyer_cols, merchant_cols = [], [], [], [], []
    for i, h in enumerate(headers):
        if _match_keywords(h, _KW_AMOUNT):
            amount_cols.append(i)
        if _match_keywords(h, _KW_INVOICE_NO):
            invoice_no_cols.append(i)
        if _match_keywords(h, _KW_DATE):
            date_cols.append(i)
        if _match_keywords(h, _KW_BUYER):
            buyer_cols.append(i)
        if _match_keywords(h, _KW_MERCHANT):
            merchant_cols.append(i)

    ambiguous = len(amount_cols) > 1 or len(invoice_no_cols) > 1

    def _first(cols):
        return cols[0] if cols else None

    return {
        "amount": _first(amount_cols),
        "invoice_no": _first(invoice_no_cols),
        "date": _first(date_cols),
        "buyer": _first(buyer_cols),
        "merchant": _first(merchant_cols),
    }, ambiguous


def _to_datetime(value):
    """把单元格值统一转成 datetime；无法解析返回 None。
    支持datetime/date对象，以及 %Y-%m-%d / %Y/%m/%d / %Y%m%d / %Y年%m月%d日 字符串。"""
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, str) and value.strip():
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y年%m月%d日"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def _buyer_match(inv_buyer, row_buyer):
    """买方模糊匹配：模板常是简写（如'南沙友谊'），库内是全称（如'广州南沙友谊人才服务有限公司'）。
    双向包含即视为匹配：任一方包含另一方。逻辑统一复用 db.buyer_match（公司归属也用同一套），避免分叉。"""
    return db.buyer_match(inv_buyer, row_buyer)


def _to_amount(value):
    """金额统一转 float；非数值返回 None。"""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


# ----------------------------------------------------------- 1. 模板解析
def parse_template(file_bytes):
    """解析 xlsx 模板字节，返回结构化信息。

    返回:
        {
            "columns": {amount/invoice_no/date/buyer/merchant: 列索引或None},  # 0-based
            "headers": [表头列表],
            "rows": [{"row_idx", "date", "merchant", "buyer", "amount", "existing_invoice_no"}, ...],
            "ambiguous": bool,
            "total_rows": int,
            "existing_invoice_no_count": int,
        }
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    columns, ambiguous = _find_columns(headers)

    col = columns  # 别名简写
    rows = []
    existing_invoice_no_count = 0
    for r in range(2, ws.max_row + 1):
        # 先判全空行
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in vals):
            continue

        def _get(idx):
            return ws.cell(row=r, column=idx + 1).value if idx is not None else None

        row_date = _to_datetime(_get(col["date"])) if col["date"] is not None else None
        merchant = _get(col["merchant"]) if col["merchant"] is not None else None
        buyer = _get(col["buyer"]) if col["buyer"] is not None else None
        amount = _to_amount(_get(col["amount"])) if col["amount"] is not None else None
        raw_no = _get(col["invoice_no"]) if col["invoice_no"] is not None else None
        existing_invoice_no = None
        if raw_no is not None and str(raw_no).strip() != "":
            existing_invoice_no = str(raw_no).strip()
            existing_invoice_no_count += 1

        rows.append({
            "row_idx": r,  # 1-based，对应原 xlsx 行号（含表头）
            "date": row_date,
            "merchant": merchant,
            "buyer": buyer,
            "amount": amount,
            "existing_invoice_no": existing_invoice_no,
        })

    wb.close()
    return {
        "columns": columns,
        "headers": headers,
        "rows": rows,
        "ambiguous": ambiguous,
        "total_rows": len(rows),
        "existing_invoice_no_count": existing_invoice_no_count,
    }


# ----------------------------------------------------------- 备注打分
def _score_candidate(remark, row_date, merchant):
    """备注模糊校验打分：含发放日期 +2，含商社名 +1。"""
    score = 0
    if remark and row_date:
        formats = [
            row_date.strftime("%Y%m%d"),
            row_date.strftime("%Y-%m-%d"),
            row_date.strftime("%Y/%m/%d"),
        ]
        if any(f in remark for f in formats):
            score += 2
    if remark and merchant and str(merchant) in remark:
        score += 1
    return score


# ----------------------------------------------------------- 2. 一对一匹配
def match_one_to_one(template_rows, candidates, config):
    """第一轮：一对一金额匹配。

    对每个待匹配行（existing_invoice_no 为空 或 overwrite=True）：
      - 按 buyer + 日期范围 [行日期, 行日期+date_range_days] 筛选候选
      - 金额精确匹配（round 到分）
      - 多候选用备注打分，取最高分
      - 已匹配发票加入 used_invoice_ids，不再复用

    返回 [{"row_idx","amount","invoice_no","invoice_id","match_type":"one_to_one","score"}, ...]
    """
    date_range_days = _cfg_get(config, "date_range_days", 30)
    overwrite = _cfg_get(config, "overwrite", False)

    used_invoice_ids = set()
    matched = []
    for row in template_rows:
        # 已有发票号且不覆盖 → 跳过
        if row.get("existing_invoice_no") and not overwrite:
            continue
        row_date = _to_datetime(row.get("date"))
        buyer = row.get("buyer")
        amount = row.get("amount")
        if row_date is None or amount is None:
            continue

        date_start = row_date
        date_end = row_date + timedelta(days=date_range_days)
        target_amount = round(amount, 2)

        avail = []
        for inv in candidates:
            if inv.get("id") in used_invoice_ids:
                continue
            if not _buyer_match(inv.get("buyer"), buyer):
                continue
            # 库内 invoice_date 可能是"2026年07月10日"等中文格式，统一解析成 date 比较
            inv_date = _to_datetime(inv.get("invoice_date"))
            if inv_date is None or inv_date < date_start or inv_date > date_end:
                continue
            if round(inv.get("amount", 0) or 0, 2) != target_amount:
                continue
            avail.append(inv)

        if not avail:
            continue

        scored = [
            (_score_candidate(inv.get("remark"), row_date, row.get("merchant")), inv)
            for inv in avail
        ]
        max_score = max(s for s, _ in scored)
        # 取最高分；并列时取首个（按候选顺序稳定）
        chosen = next(inv for s, inv in scored if s == max_score)
        used_invoice_ids.add(chosen["id"])

        matched.append({
            "row_idx": row["row_idx"],
            "amount": target_amount,
            "invoice_no": chosen["invoice_no"],
            "invoice_id": chosen["id"],
            "match_type": "one_to_one",
            "score": max_score,
        })
    return matched


# ----------------------------------------------------------- 3. 多对一凑票
def match_many_to_one(unmatched_rows, remaining_candidates, config):
    """第二轮：多对一凑票。同一买方下，多行金额相加等于一张发票金额。

    - 按 buyer 分组
    - 日期范围：组内最早发放日期 ~ 组内最晚发放日期 + date_range_days
    - 子集和：找 2~max_combo_size 行的组合，round 后等于某张发票金额
    - 单组时间预算 group_time_budget（默认 2 秒），超时跳过

    返回 [{"row_idx","amount","invoice_no","invoice_id","match_type":"many_to_one","group_id"}, ...]
    """
    date_range_days = _cfg_get(config, "date_range_days", 30)
    max_combo_size = _cfg_get(config, "max_combo_size", 5)
    group_time_budget = _cfg_get(config, "group_time_budget", 2.0)

    # 按 buyer 分组
    groups = {}
    for row in unmatched_rows:
        groups.setdefault(row.get("buyer"), []).append(row)

    matched = []
    group_counter = 0
    for buyer, rows in groups.items():
        dates = [_to_datetime(r.get("date")) for r in rows if _to_datetime(r.get("date"))]
        if not dates:
            continue
        date_start = min(dates)
        date_end = max(dates) + timedelta(days=date_range_days)

        # 该买方下日期范围内的候选发票
        buyer_invoices = []
        for inv in remaining_candidates:
            if not _buyer_match(inv.get("buyer"), buyer):
                continue
            inv_date = _to_datetime(inv.get("invoice_date"))
            if inv_date is None or inv_date < date_start or inv_date > date_end:
                continue
            buyer_invoices.append(inv)
        if not buyer_invoices:
            continue

        # 金额 → 发票列表（round 到分，避免浮点误差）
        amount_to_inv = {}
        for inv in buyer_invoices:
            amount_to_inv.setdefault(round(inv["amount"], 2), []).append(inv)

        used_invoice_ids = set()
        available_rows = [r for r in rows if r.get("amount") is not None]
        start_t = time.time()

        for size in range(2, max_combo_size + 1):
            if time.time() - start_t > group_time_budget:
                break
            if len(available_rows) < size:
                break
            # 反复尝试当前 size 的组合，直到无新匹配
            while True:
                if time.time() - start_t > group_time_budget:
                    break
                found = None
                for combo in combinations(available_rows, size):
                    if time.time() - start_t > group_time_budget:
                        break
                    combo_sum = round(sum(r["amount"] for r in combo), 2)
                    invs = amount_to_inv.get(combo_sum, [])
                    chosen_inv = next((inv for inv in invs if inv["id"] not in used_invoice_ids), None)
                    if chosen_inv:
                        found = (combo, chosen_inv)
                        break
                if not found:
                    break
                combo, chosen_inv = found
                used_invoice_ids.add(chosen_inv["id"])
                group_counter += 1
                combo_row_ids = set(id(r) for r in combo)
                for r in combo:
                    matched.append({
                        "row_idx": r["row_idx"],
                        "amount": round(r["amount"], 2),
                        "invoice_no": chosen_inv["invoice_no"],
                        "invoice_id": chosen_inv["id"],
                        "match_type": "many_to_one",
                        "group_id": group_counter,
                    })
                available_rows = [r for r in available_rows if id(r) not in combo_row_ids]
                if len(available_rows) < size:
                    break
    return matched


# ----------------------------------------------------------- 4. 结果组装
def build_result(template_rows, matched, unmatched, template_info, overwrite=False):
    """组装最终匹配结果。

    返回:
        {
            "matched": [...],
            "unmatched": [{"row_idx","amount","reason"}, ...],
            "skipped": [{"row_idx","reason":"已有发票号"}, ...],
            "template_info": {...},
            "stats": {"matched","many_to_one_groups","unmatched","skipped"},
        }
    overwrite=True 时，已填发票号的行不再归入 skipped（它们参与了匹配，
    未匹配上的已在 unmatched 中）。
    """
    matched_row_idxs = {m["row_idx"] for m in matched}

    skipped = []
    if not overwrite:
        for row in template_rows:
            if row.get("existing_invoice_no") and row["row_idx"] not in matched_row_idxs:
                skipped.append({"row_idx": row["row_idx"], "reason": "已有发票号"})

    many_to_one_groups = {
        m["group_id"] for m in matched
        if m.get("match_type") == "many_to_one" and m.get("group_id") is not None
    }

    return {
        "matched": matched,
        "unmatched": unmatched,
        "skipped": skipped,
        "template_info": template_info,
        "stats": {
            "matched": len(matched),
            "many_to_one_groups": len(many_to_one_groups),
            "unmatched": len(unmatched),
            "skipped": len(skipped),
        },
    }


# ----------------------------------------------------------- 5. 生成回填 xlsx
def generate_filled_xlsx(template_bytes, columns_map, confirmed_matched):
    """在原模板的回填列填入发票号，返回新 xlsx 字节。

    - columns_map["invoice_no"]: 回填列的 0-based 索引
    - confirmed_matched: [{"row_idx": int, "invoice_no": str}, ...]
      row_idx 对应原 xlsx 行号（1-based，含表头）
    - 用 load_workbook 保留原格式
    """
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active
    inv_col = columns_map.get("invoice_no") if columns_map else None
    if inv_col is None:
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        return buf.getvalue()

    col_1based = inv_col + 1  # 0-based → openpyxl 1-based
    for item in confirmed_matched:
        ws.cell(row=int(item["row_idx"]), column=col_1based).value = item["invoice_no"]
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


# ----------------------------------------------------------- 6. 完整流程
def run_match(template_bytes, columns_map, config, date_range_days, overwrite):
    """完整匹配流程：解析 → 取候选 → 一对一 → 多对一 → 组装结果。

    参数:
        template_bytes: xlsx 文件字节
        columns_map: 列映射覆盖（可选，None 用自动识别）；键 amount/invoice_no/date/buyer/merchant
        config: 额外配置对象/dict（可选，用于 max_combo_size/group_time_budget）
        date_range_days: 日期范围天数（显式传入，覆盖 config）
        overwrite: 是否覆盖已有发票号（显式传入）

    返回 build_result 的结构。
    """
    cfg = MatchConfig(date_range_days=date_range_days, overwrite=overwrite)
    if config is not None:
        cfg.max_combo_size = _cfg_get(config, "max_combo_size", 5)
        cfg.group_time_budget = _cfg_get(config, "group_time_budget", 2.0)

    template_info = parse_template(template_bytes)

    # 用 columns_map 覆盖自动识别的列
    cols = dict(template_info["columns"])
    if columns_map:
        for k, v in columns_map.items():
            if v is not None:
                cols[k] = v
    template_info["columns"] = cols

    template_rows = template_info["rows"]

    # 按买方取候选发票（日期范围取该买方所有行的最早~最晚+N天）
    buyers = {r.get("buyer") for r in template_rows if r.get("buyer")}
    all_candidates = []
    for buyer in buyers:
        bdates = [r["date"] for r in template_rows
                  if r.get("buyer") == buyer and r.get("date")]
        if not bdates:
            continue
        d_from = min(bdates).strftime("%Y-%m-%d")
        d_to = (max(bdates) + timedelta(days=date_range_days)).strftime("%Y-%m-%d")
        all_candidates.extend(db.get_invoices_for_matching(buyer, d_from, d_to))

    # 第一轮：一对一
    matched = match_one_to_one(template_rows, all_candidates, cfg)
    matched_row_idxs = {m["row_idx"] for m in matched}
    used_invoice_ids = {m["invoice_id"] for m in matched}

    # 收集未匹配行 + 原因
    unmatched_rows = []
    unmatched = []
    for row in template_rows:
        if row.get("existing_invoice_no") and not overwrite:
            continue  # 跳过（已有发票号）
        if row["row_idx"] in matched_row_idxs:
            continue
        unmatched_rows.append(row)
        buyer_candidates = [c for c in all_candidates if _buyer_match(c.get("buyer"), row.get("buyer"))]
        if not buyer_candidates:
            reason = "无候选"
        else:
            reason = "金额不匹配"
        amt = round(row["amount"], 2) if row.get("amount") is not None else None
        unmatched.append({"row_idx": row["row_idx"], "amount": amt, "reason": reason})

    # 第二轮：多对一
    remaining_candidates = [c for c in all_candidates if c["id"] not in used_invoice_ids]
    m2o = match_many_to_one(unmatched_rows, remaining_candidates, cfg)
    matched.extend(m2o)
    m2o_row_idxs = {m["row_idx"] for m in m2o}
    # 从 unmatched 中移除已被多对一匹配的行
    unmatched = [u for u in unmatched if u["row_idx"] not in m2o_row_idxs]

    return build_result(template_rows, matched, unmatched, template_info, overwrite=overwrite)
