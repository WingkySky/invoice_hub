"""
命令行入口 (CLI) —— 只做"调度 + 人类可读输出"，逻辑都在 api / engine / db 里。

agent / 程序化调用建议用 --json：每条命令输出一行 JSON，便于解析，不再啃打印文本。
示例：
  python hub.py --json accounts list
  python hub.py --json accounts test 3
  python hub.py --json fetch --since 2026-07-09

用法:
  python hub.py init                 # 建库 + 从 config/accounts.json 播种账号
  python hub.py accounts list        # 列出已管理的邮箱
  python hub.py accounts add --email x@163.com --name 别名 --password 授权码 [--host imap.163.com --port 993]
  python hub.py accounts toggle <id> <0|1>
  python hub.py accounts test <id>   # 测试单账号 IMAP 连接（不进 Web 也能验证，agent 友好）
  python hub.py fetch [--since 2026-07-09]   # 拉取+解析全部启用账号
  python hub.py seed <本地PDF文件夹> [--account 别名]  # 把已有PDF导入库（演示/补录）
  python hub.py report [--xlsx 路径] [--html 路径]     # 从库导出 Excel/HTML（离线用）
  python hub.py serve [--port 8000]                   # 起通用 Web 控制台
  python hub.py companies list                        # 列出所属公司
  python hub.py companies add --name "某某科技" [--tax-id 9111...] [--aliases "简称A,简称B"]
  python hub.py companies update <id> --name "新名" [--tax-id ...] [--aliases ...]
  python hub.py companies delete <id>                 # 删除公司（名下发票回退未归类）
  python hub.py companies import                      # 从发票购买方批量导入候选公司
  python hub.py companies backfill                    # 历史发票一键回填归属
  python hub.py companies assign --ids 1,2,3 --company-id <id>   # 批量改所属公司
  （任意命令加 --json 即以 JSON 输出）
"""
import argparse
import json
import os
import sys
import glob

import db
import engine
import api

HERE = os.path.dirname(os.path.abspath(__file__))
ACCOUNTS_SEED = os.path.join(HERE, "config", "accounts.json")

# --json 模式开关：置 True 时所有命令输出一行 JSON，且静默抓取过程日志（engine.QUIET）。
OUT_JSON = False


def _result(data, human_lines):
    """统一输出：--json 时打印 JSON，否则逐行打印人类可读文本。

    data        : 可 JSON 序列化的 dict（最终结构）
    human_lines : 人类可读文本列表（每行一项）
    """
    if OUT_JSON:
        print(json.dumps(data, ensure_ascii=False))
    else:
        for line in human_lines:
            print(line)


# ----------------------------------------------------------- init
def cmd_init():
    path = db.init()
    seeded = 0
    if os.path.exists(ACCOUNTS_SEED):
        with open(ACCOUNTS_SEED, encoding="utf-8") as f:
            cfg = json.load(f)
        for a in cfg.get("accounts", []):
            db.upsert_account({
                "name": a.get("name", a["email"]),
                "email": a["email"],
                "provider": a.get("provider"),
                "imap_host": a.get("imap_host"),
                "imap_port": a.get("imap_port", 993),
                "use_ssl": 1 if a.get("use_ssl", True) else 0,
                "folder": a.get("folder", "INBOX"),
                "password": a.get("password"),
                "enabled": 1 if a.get("enabled", True) else 0,
            })
            seeded += 1
        human = [f"[ok] 数据库已初始化: {path}",
                 f"[ok] 已从 config/accounts.json 播种 {seeded} 个账号（授权码请在库里用 accounts add 修正）"]
    else:
        human = [f"[ok] 数据库已初始化: {path}",
                 "[提示] 未找到 config/accounts.json，跳过账号播种。用 `hub.py accounts add` 手动加。"]
    _result({"ok": True, "db": path, "seeded": seeded}, human)


# ----------------------------------------------------------- accounts
def cmd_accounts(args):
    if args.sub == "list" or args.sub is None:
        rows = api.list_accounts()  # 已脱敏（不含明文密码）
        human = []
        if not rows:
            human.append("（暂无账号）用 `hub.py accounts add` 添加。")
        else:
            human.append(f"{'ID':<3} {'启用':<4} {'名称':<14} {'邮箱':<28} {'IMAP':<22} 最近抓取")
            for a in rows:
                human.append(f"{a['id']:<3} {'✓' if a['enabled'] else '✗':<4} {str(a['name'])[:13]:<14} "
                             f"{a['email']:<28} {str(a['imap_host'])+':'+str(a['imap_port']):<22} {a['last_fetch'] or '-'}")
        _result({"ok": True, "accounts": rows}, human)
        return

    if args.sub == "add":
        acc = api.add_account({
            "name": args.name or args.email,
            "email": args.email,
            "provider": args.provider,
            "imap_host": args.host or "imap.163.com",
            "imap_port": args.port or 993,
            "use_ssl": 1,
            "folder": args.folder or "INBOX",
            "password": args.password,
            "enabled": 1,
        })
        _result(acc, [f"[ok] 已添加/更新账号 {args.email}"])
        return

    if args.sub == "toggle":
        r = api.toggle_account(args.id, bool(args.enabled))
        _result(r, [f"[ok] 账号 #{args.id} 已{'启用' if args.enabled else '停用'}"])
        return

    if args.sub == "test":
        r = api.test_connection(args.id)
        prefix = "✓ " if r["ok"] else "✗ "
        _result(r, [prefix + r["msg"]])
        return


# ----------------------------------------------------------- companies（公司归属维度）
def cmd_companies(args):
    if args.sub == "list" or args.sub is None:
        rows = api.list_companies()
        human = []
        if not rows:
            human.append("（暂无公司）用 `hub.py companies add` 添加，或 `hub.py companies import` 从发票导入。")
        else:
            human.append(f"{'ID':<3} {'名称':<22} {'税号':<22} 别名")
            for c in rows:
                try:
                    al = c.get("aliases") or "[]"
                    al = json.loads(al) if isinstance(al, str) else (al or [])
                except Exception:
                    al = []
                al_s = "、".join(al) if isinstance(al, list) else str(al)
                human.append(f"{c['id']:<3} {str(c['name'])[:21]:<22} {str(c.get('tax_id') or '-')[:21]:<22} {al_s}")
        _result({"ok": True, "companies": rows}, human)
        return

    if args.sub == "add":
        if not args.name:
            _result({"ok": False, "error": "name_required"}, ["[错误] --name 必填"])
            sys.exit(1)
        c = api.create_company({"name": args.name, "tax_id": args.tax_id or "", "aliases": args.aliases or ""})
        _result({"ok": True, "company": c}, [f"[ok] 已添加公司 {c.get('name')} (id={c.get('id')})"])

    if args.sub == "update":
        fields = {}
        if args.name is not None:
            fields["name"] = args.name
        if args.tax_id is not None:
            fields["tax_id"] = args.tax_id
        if args.aliases is not None:
            fields["aliases"] = args.aliases
        c = api.update_company(args.id, fields)
        _result({"ok": True, "company": c}, [f"[ok] 已更新公司 #{args.id} -> {c.get('name')}"])

    if args.sub == "delete":
        r = api.delete_company(args.id)
        _result(r, [f"[ok] 已删除公司 #{args.id}（其名下发票回退为未归类）"])

    if args.sub == "import":
        r = api.import_companies_from_invoices()
        _result(r, [f"[ok] 已从发票导入 {r.get('added', 0)} 个候选公司"])

    if args.sub == "backfill":
        r = api.backfill_company()
        human = [f"[ok] 回填完成：扫描 {r.get('scanned', 0)} 张",
                 f"     已归类 {r.get('classified', 0)} / 未归类 {r.get('unclassified', 0)} / 歧义 {r.get('ambiguous', 0)}"]
        _result(r, human)

    if args.sub == "assign":
        ids = [int(x) for x in (args.ids or "").split(",") if x.strip().isdigit()]
        if not ids:
            _result({"ok": False, "error": "ids_required"}, ["[错误] --ids 必填（逗号分隔的发票 id）"])
            sys.exit(1)
        cid = int(args.company_id) if args.company_id else None
        r = api.assign_invoices(ids, cid)
        human = [f"[ok] 已将 {r.get('updated', 0)} 张发票归属到"
                 + (f" 公司 #{cid}" if cid else " 未归类")]
        _result(r, human)


# ----------------------------------------------------------- fetch
def cmd_fetch(args):
    db.init()
    # fetch_all 内部用 engine.log 打印进度：
    #   - 人类模式：实时打印过程日志（含"本次新增 X 张"）
    #   - --json 模式：engine.QUIET=True 已静默日志，这里只输出最终 JSON
    new = engine.fetch_all(since_override=args.since)
    if OUT_JSON:
        print(json.dumps({"ok": True, "new": int(new), "since": args.since}, ensure_ascii=False))


# ----------------------------------------------------------- seed (本地PDF入库)
def cmd_seed(args):
    db.init()
    folder = args.folder
    if not os.path.isdir(folder):
        _result({"ok": False, "error": "folder_not_found", "folder": folder},
                [f"[错误] 找不到文件夹 {folder}"])
        sys.exit(1)
    # 找/建一个"本地"账号
    acc = None
    for a in db.get_accounts():
        if a["email"] == "local@seed":
            acc = a
            break
    if not acc:
        db.upsert_account({"name": args.account or "本地导入", "email": "local@seed",
                            "provider": "seed", "imap_host": "", "imap_port": 0,
                            "use_ssl": 0, "folder": "", "password": "", "enabled": 1})
        acc = db.get_accounts()[-1]
    pdfs = glob.glob(os.path.join(folder, "**", "*.pdf"), recursive=True)
    new = 0
    lines = [f"找到 {len(pdfs)} 个 PDF，开始解析入库..."]
    for p in pdfs:
        inv = engine.parse_pdf_to_invoice(p, acc["id"], email_id=None, source_type="seed")
        if db.insert_invoice(inv):
            new += 1
            lines.append(f"  + {os.path.basename(p)} -> 买方={inv['buyer'] or '?'} 金额={inv['amount']} 号={inv['invoice_no'] or '?'}")
    lines.append(f"[ok] 新增 {new} 张（共 {len(pdfs)} 个PDF）")
    _result({"ok": True, "added": new, "total": len(pdfs)}, lines)


# ----------------------------------------------------------- report (离线导出)
def cmd_report(args):
    rows = db.get_invoices()
    if not rows:
        _result({"ok": False, "error": "empty"}, ["（库里还没有发票）先 fetch 或 seed。"])
        return
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    xlsx = args.xlsx or os.path.join(HERE, "data", "发票汇总.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "发票记录"
    headers = ["#", "邮箱", "买方主体", "发票号码", "金额", "开票日期", "销售方", "城市", "PDF"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="534AB7")
        c.alignment = Alignment(horizontal="center")
    total = 0.0
    for i, r in enumerate(rows, 1):
        try:
            total += float(r.get("amount") or 0)
        except Exception:
            pass
        ws.append([i, r.get("account_name") or r.get("account_email") or "", r.get("buyer", ""),
                   r.get("invoice_no", ""), r.get("amount", ""), r.get("invoice_date", ""),
                   r.get("seller", ""), r.get("city", ""), r.get("pdf_path", "")])
    ws.append(["", "", "", "", "合计", round(total, 2), "", "", ""])
    for col, w in zip("ABCDEFGHI", [4, 18, 26, 22, 12, 14, 26, 10, 30]):
        ws.column_dimensions[col].width = w
    os.makedirs(os.path.dirname(xlsx), exist_ok=True)
    wb.save(xlsx)

    html = args.html or os.path.join(HERE, "data", "发票汇总.html")
    trs = ""
    for i, r in enumerate(rows, 1):
        trs += (f"<tr><td>{i}</td><td>{r.get('account_name','')}</td><td>{r.get('buyer','')}</td>"
                f"<td>{r.get('invoice_no','')}</td><td style='text-align:right'>{r.get('amount','')}</td>"
                f"<td>{r.get('invoice_date','')}</td><td>{r.get('seller','')}</td>"
                f"<td>{r.get('city','')}</td><td>{r.get('pdf_path','')}</td></tr>")
    with open(html, "w", encoding="utf-8") as f:
        f.write(f"""<!doctype html><html lang="zh"><head><meta charset="utf-8">
<title>发票汇总</title><style>
body{{font-family:system-ui,'Microsoft YaHei';margin:24px;color:#222}}
h1{{font-size:18px}} .sum{{font-size:15px;margin:8px 0 16px}}
table{{border-collapse:collapse;width:100%;font-size:13px}}
th,td{{border:1px solid #ddd;padding:6px 8px}}
th{{background:#534AB7;color:#fff}} tr:nth-child(even){{background:#f7f6fb}}
</style></head><body>
<h1>发票汇总（{len(rows)} 张）</h1>
<div class="sum">合计金额：<b>¥{round(total,2)}</b></div>
<table><thead><tr><th>#</th><th>邮箱</th><th>买方主体</th><th>发票号码</th><th>金额</th>
<th>开票日期</th><th>销售方</th><th>城市</th><th>PDF</th></tr></thead>
<tbody>{trs}</tbody></table></body></html>""")
    _result({"ok": True, "xlsx": xlsx, "html": html, "count": len(rows), "total": round(total, 2)},
            [f"[ok] Excel: {xlsx}", f"[ok] HTML : {html}", f"合计 ¥{round(total,2)}，{len(rows)} 张"])


# ----------------------------------------------------------- serve
def cmd_serve(args):
    from web import app
    if OUT_JSON:
        print(json.dumps({"ok": True, "mode": "serve", "port": args.port or 8000}, ensure_ascii=False))
    app.run(port=args.port or 8000)


# ----------------------------------------------------------- 参数解析
def build_parser():
    ap = argparse.ArgumentParser(prog="hub.py", description="票归集 · 发票归集中枢（数据驱动）")
    ap.add_argument("--json", action="store_true",
                    help="以 JSON 输出结果（每行一个 JSON 对象），便于 agent / 程序解析")
    sub = ap.add_subparsers(dest="cmd")

    sub.add_parser("init", help="建库 + 播种账号")

    p_acc = sub.add_parser("accounts", help="管理邮箱账号")
    p_acc_sub = p_acc.add_subparsers(dest="sub")
    p_acc_sub.add_parser("list")
    p_acc_add = p_acc_sub.add_parser("add")
    p_acc_add.add_argument("--email", required=True)
    p_acc_add.add_argument("--name")
    p_acc_add.add_argument("--password", required=True)
    p_acc_add.add_argument("--host")
    p_acc_add.add_argument("--port", type=int)
    p_acc_add.add_argument("--provider")
    p_acc_add.add_argument("--folder")
    p_acc_tog = p_acc_sub.add_parser("toggle")
    p_acc_tog.add_argument("id", type=int)
    p_acc_tog.add_argument("enabled", type=int, choices=[0, 1])
    p_acc_test = p_acc_sub.add_parser("test", help="测试单账号 IMAP 连接")
    p_acc_test.add_argument("id", type=int)

    p_fetch = sub.add_parser("fetch", help="拉取+解析全部启用账号")
    p_fetch.add_argument("--since", default=None,
                         help="临时覆盖抓取起始日期（如 2026-07-01 或 90d）；留空用各账号偏好")

    p_seed = sub.add_parser("seed", help="把本地PDF文件夹导入库")
    p_seed.add_argument("folder")
    p_seed.add_argument("--account")

    p_rep = sub.add_parser("report", help="从库导出 Excel/HTML")
    p_rep.add_argument("--xlsx")
    p_rep.add_argument("--html")

    p_serve = sub.add_parser("serve", help="起 Web 控制台")
    p_serve.add_argument("--port", type=int, default=8000)

    p_comp = sub.add_parser("companies", help="管理所属公司（发票归集维度）")
    p_comp_sub = p_comp.add_subparsers(dest="sub")
    p_comp_sub.add_parser("list")
    p_comp_add = p_comp_sub.add_parser("add")
    p_comp_add.add_argument("--name", help="公司名称（必填）")
    p_comp_add.add_argument("--tax-id", dest="tax_id", help="统一社会信用代码（18位，可留空自动回填）")
    p_comp_add.add_argument("--aliases", help="识别别名，逗号分隔（用于匹配购买方）")
    p_comp_upd = p_comp_sub.add_parser("update")
    p_comp_upd.add_argument("id", type=int)
    p_comp_upd.add_argument("--name")
    p_comp_upd.add_argument("--tax-id", dest="tax_id")
    p_comp_upd.add_argument("--aliases")
    p_comp_del = p_comp_sub.add_parser("delete")
    p_comp_del.add_argument("id", type=int)
    p_comp_sub.add_parser("import").help = "从已归集发票的购买方批量导入候选公司"
    p_comp_sub.add_parser("backfill").help = "对历史未归类/歧义发票重跑归属逻辑"
    p_comp_assign = p_comp_sub.add_parser("assign")
    p_comp_assign.add_argument("--ids", required=True, help="逗号分隔的发票 id")
    p_comp_assign.add_argument("--company-id", dest="company_id", help="目标公司 id（省略=置为未归类）")

    return ap


# ----------------------------------------------------------- main
def main():
    ap = build_parser()
    args = ap.parse_args()

    # 全局开关：--json 时静默抓取日志，只输出最终 JSON
    global OUT_JSON
    OUT_JSON = args.json
    if OUT_JSON:
        engine.QUIET = True

    if args.cmd == "init":
        cmd_init()
    elif args.cmd == "accounts":
        cmd_accounts(args)
    elif args.cmd == "fetch":
        cmd_fetch(args)
    elif args.cmd == "seed":
        cmd_seed(args)
    elif args.cmd == "report":
        cmd_report(args)
    elif args.cmd == "serve":
        cmd_serve(args)
    elif args.cmd == "companies":
        cmd_companies(args)
    else:
        ap.print_help()


if __name__ == "__main__":
    main()
