"""
matching.py 单元测试。

覆盖：
  1. parse_template 标准模板列识别
  2. parse_template 多金额列 → ambiguous=True
  3. match_one_to_one 金额唯一匹配成功
  4. match_one_to_one 多候选用备注打分选最优
  5. match_one_to_one 已匹配发票不重复使用
  6. match_many_to_one 两行凑一张发票成功
  7. match_many_to_one 组合超 5 行不尝试
  8. generate_filled_xlsx 正确回填发票号
"""
import io
import os
import sys
import unittest
from datetime import datetime

from openpyxl import Workbook, load_workbook

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if HERE not in sys.path:
    sys.path.insert(0, HERE)

import matching
from matching import (
    MatchConfig,
    parse_template,
    match_one_to_one,
    match_many_to_one,
    generate_filled_xlsx,
)


def _make_xlsx(headers, rows):
    """用 openpyxl 造一个 xlsx 字节流。headers 是表头列表，rows 是每行值列表。"""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _inv(id_, buyer, amount, invoice_no, invoice_date, remark=""):
    """造一个候选发票 dict（与 db.get_invoices_for_matching 返回结构一致）。"""
    return {
        "id": id_, "buyer": buyer, "amount": amount, "invoice_no": invoice_no,
        "invoice_date": invoice_date, "remark": remark, "note": "",
        "seller": "", "pdf_path": "",
    }


class ParseTemplateTest(unittest.TestCase):
    def test_parse_template_standard_columns(self):
        """标准 8 列模板能自动识别 5 类列，且 ambiguous=False。"""
        headers = ["发放日期", "所属部门", "商社", "收款单位", "支出金额",
                   "平台", "渠道", "平台开票号"]
        rows = [
            [datetime(2026, 6, 30), "外包A部", "晟联数码", "南沙友谊", 36925.01,
             "身边云", "博跃", "26122000000877888861"],
            [datetime(2026, 7, 10), "外包B部", "另一商社", "南沙友谊", 949.50,
             "身边云", "博跃", None],
        ]
        info = parse_template(_make_xlsx(headers, rows))

        self.assertEqual(info["columns"]["date"], 0)
        self.assertEqual(info["columns"]["merchant"], 2)
        self.assertEqual(info["columns"]["buyer"], 3)
        self.assertEqual(info["columns"]["amount"], 4)
        self.assertEqual(info["columns"]["invoice_no"], 7)
        self.assertFalse(info["ambiguous"])

        self.assertEqual(info["total_rows"], 2)
        self.assertEqual(info["existing_invoice_no_count"], 1)

        r0 = info["rows"][0]
        self.assertEqual(r0["row_idx"], 2)  # 1-based，含表头
        self.assertEqual(r0["buyer"], "南沙友谊")
        self.assertEqual(r0["merchant"], "晟联数码")
        self.assertEqual(r0["amount"], 36925.01)
        self.assertIsNotNone(r0["existing_invoice_no"])
        self.assertEqual(r0["date"], datetime(2026, 6, 30))

        r1 = info["rows"][1]
        self.assertEqual(r1["amount"], 949.50)
        self.assertIsNone(r1["existing_invoice_no"])

    def test_parse_template_ambiguous_multiple_amount_cols(self):
        """多个金额列 → ambiguous=True。"""
        headers = ["发放日期", "商社", "收款单位", "金额", "支出金额", "平台开票号"]
        rows = [[datetime(2026, 7, 10), "商A", "买方A", 100.0, 200.0, None]]
        info = parse_template(_make_xlsx(headers, rows))
        self.assertTrue(info["ambiguous"])

    def test_parse_template_skips_empty_rows(self):
        """全空行被跳过，不计入 total_rows。"""
        headers = ["发放日期", "商社", "收款单位", "支出金额", "平台开票号"]
        rows = [
            [datetime(2026, 7, 10), "商A", "买方A", 100.0, None],
            [None, None, None, None, None],  # 全空行
            [datetime(2026, 7, 11), "商B", "买方A", 200.0, None],
        ]
        info = parse_template(_make_xlsx(headers, rows))
        self.assertEqual(info["total_rows"], 2)


class MatchOneToOneTest(unittest.TestCase):
    def test_unique_amount_match(self):
        """金额唯一匹配成功。"""
        rows = [{"row_idx": 2, "date": datetime(2026, 7, 1), "merchant": "商A",
                 "buyer": "买方A", "amount": 949.50, "existing_invoice_no": None}]
        candidates = [_inv(1, "买方A", 949.50, "INV001", "2026-07-05")]
        matched = match_one_to_one(rows, candidates, MatchConfig())
        self.assertEqual(len(matched), 1)
        self.assertEqual(matched[0]["invoice_no"], "INV001")
        self.assertEqual(matched[0]["invoice_id"], 1)
        self.assertEqual(matched[0]["match_type"], "one_to_one")

    def test_multi_candidate_remark_scoring(self):
        """多候选时用备注打分（日期+2、商社+1）选最优。"""
        rows = [{"row_idx": 2, "date": datetime(2026, 6, 30), "merchant": "晟联数码",
                 "buyer": "南沙友谊", "amount": 36925.01, "existing_invoice_no": None}]
        candidates = [
            _inv(1, "南沙友谊", 36925.01, "INV_A", "2026-06-30",
                 remark="发放日期20260630 晟联数码"),
            _inv(2, "南沙友谊", 36925.01, "INV_B", "2026-06-30",
                 remark="其他内容"),
        ]
        matched = match_one_to_one(rows, candidates, MatchConfig())
        self.assertEqual(len(matched), 1)
        # INV_A 含日期(+2) + 商社(+1) = 3 分，优于 INV_B 的 0 分
        self.assertEqual(matched[0]["invoice_no"], "INV_A")
        self.assertEqual(matched[0]["score"], 3)

    def test_matched_invoice_not_reused(self):
        """已匹配的发票不再参与后续行的匹配。"""
        rows = [
            {"row_idx": 2, "date": datetime(2026, 7, 1), "merchant": "商A",
             "buyer": "买方A", "amount": 100.0, "existing_invoice_no": None},
            {"row_idx": 3, "date": datetime(2026, 7, 1), "merchant": "商A",
             "buyer": "买方A", "amount": 100.0, "existing_invoice_no": None},
        ]
        # 仅一张 100 元发票，两行都要它 → 只能匹配第一行
        candidates = [_inv(1, "买方A", 100.0, "INV001", "2026-07-05")]
        matched = match_one_to_one(rows, candidates, MatchConfig())
        self.assertEqual(len(matched), 1)
        self.assertEqual(matched[0]["row_idx"], 2)
        self.assertEqual(matched[0]["invoice_no"], "INV001")

    def test_skip_existing_invoice_no_without_overwrite(self):
        """已有发票号且未开启覆盖 → 跳过，不参与匹配。"""
        rows = [{"row_idx": 2, "date": datetime(2026, 7, 1), "merchant": "商A",
                 "buyer": "买方A", "amount": 100.0, "existing_invoice_no": "EXISTING"}]
        candidates = [_inv(1, "买方A", 100.0, "INV001", "2026-07-05")]
        # 默认不覆盖
        self.assertEqual(match_one_to_one(rows, candidates, MatchConfig()), [])
        # 开启覆盖后匹配
        matched = match_one_to_one(rows, candidates, MatchConfig(overwrite=True))
        self.assertEqual(len(matched), 1)
        self.assertEqual(matched[0]["invoice_no"], "INV001")

    def test_date_range_filtering(self):
        """超出日期范围的候选不参与匹配。"""
        rows = [{"row_idx": 2, "date": datetime(2026, 7, 1), "merchant": "商A",
                 "buyer": "买方A", "amount": 100.0, "existing_invoice_no": None}]
        # 发票日期 2026-08-15，距发放日 45 天，超过默认 30 天范围
        candidates = [_inv(1, "买方A", 100.0, "INV001", "2026-08-15")]
        self.assertEqual(match_one_to_one(rows, candidates, MatchConfig()), [])
        # 放宽到 60 天后可匹配
        matched = match_one_to_one(rows, candidates, MatchConfig(date_range_days=60))
        self.assertEqual(len(matched), 1)


class MatchManyToOneTest(unittest.TestCase):
    def test_two_rows_sum_to_one_invoice(self):
        """两行金额相加等于一张发票 → 两行均回填同一发票号。"""
        rows = [
            {"row_idx": 2, "date": datetime(2026, 7, 1), "merchant": "商A",
             "buyer": "买方A", "amount": 100.0, "existing_invoice_no": None},
            {"row_idx": 3, "date": datetime(2026, 7, 1), "merchant": "商A",
             "buyer": "买方A", "amount": 200.0, "existing_invoice_no": None},
        ]
        candidates = [_inv(1, "买方A", 300.0, "INV300", "2026-07-05")]
        matched = match_many_to_one(rows, candidates, MatchConfig())
        self.assertEqual(len(matched), 2)
        self.assertEqual({m["invoice_no"] for m in matched}, {"INV300"})
        self.assertEqual({m["match_type"] for m in matched}, {"many_to_one"})
        # 同一组 group_id
        gids = {m["group_id"] for m in matched}
        self.assertEqual(len(gids), 1)

    def test_combo_over_limit_not_attempted(self):
        """需要 6 行才能凑成一张发票时，max_combo_size=5 不尝试 → 全部未匹配。"""
        rows = [
            {"row_idx": i + 2, "date": datetime(2026, 7, 1), "merchant": "商A",
             "buyer": "买方A", "amount": 50.0, "existing_invoice_no": None}
            for i in range(6)
        ]
        candidates = [_inv(1, "买方A", 300.0, "INV300", "2026-07-05")]
        matched = match_many_to_one(rows, candidates, MatchConfig(max_combo_size=5))
        self.assertEqual(len(matched), 0)

    def test_smaller_combo_preferred(self):
        """2 行能凑时优先用 2 行，而非更多行。"""
        rows = [
            {"row_idx": 2, "date": datetime(2026, 7, 1), "merchant": "商A",
             "buyer": "买方A", "amount": 100.0, "existing_invoice_no": None},
            {"row_idx": 3, "date": datetime(2026, 7, 1), "merchant": "商A",
             "buyer": "买方A", "amount": 100.0, "existing_invoice_no": None},
            {"row_idx": 4, "date": datetime(2026, 7, 1), "merchant": "商A",
             "buyer": "买方A", "amount": 100.0, "existing_invoice_no": None},
            {"row_idx": 5, "date": datetime(2026, 7, 1), "merchant": "商A",
             "buyer": "买方A", "amount": 100.0, "existing_invoice_no": None},
        ]
        # 一张 200 元发票：2 行 100 凑成，剩 2 行无对应发票
        candidates = [_inv(1, "买方A", 200.0, "INV200", "2026-07-05")]
        matched = match_many_to_one(rows, candidates, MatchConfig())
        self.assertEqual(len(matched), 2)
        self.assertEqual({m["invoice_no"] for m in matched}, {"INV200"})


class GenerateFilledXlsxTest(unittest.TestCase):
    def test_fill_invoice_no(self):
        """回填发票号到指定列，保留其他单元格。"""
        headers = ["发放日期", "商社", "收款单位", "支出金额", "平台开票号"]
        rows = [
            [datetime(2026, 7, 1), "商A", "买方A", 100.0, None],
            [datetime(2026, 7, 2), "商B", "买方A", 200.0, None],
        ]
        data = _make_xlsx(headers, rows)
        columns_map = {"invoice_no": 4}  # 0-based：平台开票号是第 5 列
        confirmed = [
            {"row_idx": 2, "invoice_no": "INV001"},
            {"row_idx": 3, "invoice_no": "INV002"},
        ]
        out = generate_filled_xlsx(data, columns_map, confirmed)

        wb = load_workbook(io.BytesIO(out))
        ws = wb.active
        # 0-based col 4 → 1-based col 5
        self.assertEqual(ws.cell(row=2, column=5).value, "INV001")
        self.assertEqual(ws.cell(row=3, column=5).value, "INV002")
        # 其他列保留
        self.assertEqual(ws.cell(row=2, column=4).value, 100.0)
        self.assertEqual(ws.cell(row=1, column=5).value, "平台开票号")
        wb.close()

    def test_preserves_existing_invoice_no(self):
        """回填不覆盖已有发票号的单元格（仅写 confirmed 指定的行）。"""
        headers = ["支出金额", "平台开票号"]
        rows = [
            [100.0, "KEEP_ME"],
            [200.0, None],
        ]
        data = _make_xlsx(headers, rows)
        columns_map = {"invoice_no": 1}
        confirmed = [{"row_idx": 3, "invoice_no": "INV002"}]  # 只回填第 3 行
        out = generate_filled_xlsx(data, columns_map, confirmed)

        wb = load_workbook(io.BytesIO(out))
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=2).value, "KEEP_ME")  # 未触碰
        self.assertEqual(ws.cell(row=3, column=2).value, "INV002")
        wb.close()


class BuildResultTest(unittest.TestCase):
    """build_result 基本组装校验（非核心，确保 stats/skipped 正确）。"""

    def test_stats_and_skipped(self):
        template_rows = [
            {"row_idx": 2, "amount": 100.0, "existing_invoice_no": "OLD"},
            {"row_idx": 3, "amount": 200.0, "existing_invoice_no": None},
            {"row_idx": 4, "amount": 300.0, "existing_invoice_no": None},
        ]
        matched = [
            {"row_idx": 3, "amount": 200.0, "invoice_no": "INV2",
             "invoice_id": 2, "match_type": "one_to_one", "score": 0},
            {"row_idx": 4, "amount": 150.0, "invoice_no": "INV3",
             "invoice_id": 3, "match_type": "many_to_one", "group_id": 1},
        ]
        unmatched = [{"row_idx": 99, "amount": 0.0, "reason": "无候选"}]
        result = matching.build_result(template_rows, matched, unmatched, {"rows": 3})

        self.assertEqual(result["stats"]["matched"], 2)
        self.assertEqual(result["stats"]["many_to_one_groups"], 1)
        self.assertEqual(result["stats"]["unmatched"], 1)
        self.assertEqual(result["stats"]["skipped"], 1)  # row 2 已有发票号
        self.assertEqual(result["skipped"][0]["row_idx"], 2)


if __name__ == "__main__":
    unittest.main()
